from etherscan import Etherscan
from openpyxl import Workbook, load_workbook
import csv
from datetime import datetime
from variable_dic import *
import sys


YOUR_API_KEY = '7HZJFYIG6XT58SP3VWU49NJJ3FSRHP23H6980'

eth = Etherscan(YOUR_API_KEY)


def write_daily_account_balance(day, ws_new, name):
    row_index = chr(64 + day)

    cnt = 1

    for row_cells, row_cells1 in zip(ws_new.iter_cols(min_row=day, max_row=day),  ws_new.iter_cols(min_row=1, max_row=1)):
        cnt += 1
        for cell, cell1 in zip(row_cells, row_cells1):
            print(cnt)
            if cnt < 100:
                cell.value = eth.get_acc_balance_by_token_and_contract_address(name, cell1.value)   #change here



def main():
    index = 0
    ok = 0
    for key, value in dic_var.items():
        index = index + 1
        if 'inch' in key:
            ok = 1
        if 'holders' in key and ok == 1:
            print(value)
            day_of_the_year = datetime.now().timetuple().tm_yday - 10

            wb_new = load_workbook(value)
            ws_new = wb_new.active

            name = list(dic_var.items())[index - 2][1]

            write_daily_account_balance(day_of_the_year, ws_new, name)

            wb_new.save(value)




if __name__ == "__main__":
    main()




